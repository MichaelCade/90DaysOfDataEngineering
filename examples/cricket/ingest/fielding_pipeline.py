"""Cricket applied example — ingest Uffington 1st XI fielding stats with dlt.

The third source, completing batting + bowling + fielding. Like bowling, the fielding stats
table has a built-in **Export to CSV** on Play-Cricket, so acquisition is just: export, drop the
.xlsx in data/snapshots/, run. Same governed `merge` load as the others.

Fielding covers both wicket-keeping (catches, stumpings) and out-fielding (catches, run-outs),
so the export has a wider set of columns — dlt infers them all.

    pip install openpyxl
    export DESTINATION__POSTGRES__CREDENTIALS="postgresql://app:<pw>@192.168.169.191:5432/appdb"
    python fielding_pipeline.py                        # loads fielding_2026_v1.xlsx
    python fielding_pipeline.py fielding_2026_v2.xlsx  # a later export -> merges on top

Data lands in the `cricket` schema of `appdb` (table: fielding).
"""
import os
import sys
import dlt
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SNAP_DIR = os.environ.get("CRICKET_SNAPSHOT_DIR", os.path.join(HERE, "..", "data", "snapshots"))
SNAPSHOT = (sys.argv[1] if len(sys.argv) > 1
            else os.environ.get("CRICKET_FIELDING_SNAPSHOT", "fielding_2026_v1.xlsx"))
SEASON = int(os.environ.get("CRICKET_SEASON", "2026"))


@dlt.resource(name="fielding", write_disposition="merge", primary_key=("player", "season"))
def fielding(snapshot: str):
    """Read one fielding export (xlsx) and yield a row per player (header row -> dict keys).

    dlt normalizes the headers to safe column names ("WICKET KEEPING CATCHES" ->
    wicket_keeping_catches, "RUN OUTS" -> run_outs) and infers types.
    """
    wb = openpyxl.load_workbook(os.path.join(SNAP_DIR, snapshot), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    for values in rows[1:]:
        if not values or values[1] in (None, "", "Player"):   # skip blanks; col 1 = Player
            continue
        record = dict(zip(header, values))
        record["season"] = SEASON
        yield record


def main() -> None:
    pipeline = dlt.pipeline(pipeline_name="cricket", destination="postgres", dataset_name="cricket")
    load_info = pipeline.run(fielding(SNAPSHOT))
    print(f"loaded snapshot: {SNAPSHOT}")
    print(load_info)
    print(f"  fielding: {len(pipeline.dataset()['fielding'].fetchall())} rows total in cricket.fielding")


if __name__ == "__main__":
    main()
