"""Cricket applied example — ingest Uffington 1st XI bowling stats with dlt.

Companion to batting_pipeline.py. Unlike the batting run-ranges (which need a browser scrape),
Play-Cricket's bowling statistics table has a built-in **Export to CSV** — so that's the
acquisition path here: click export, drop the .xlsx into data/snapshots/. Cleaner and more
robust than scraping.

dlt doesn't care that the source is a spreadsheet — we read the sheet with openpyxl and yield
one dict per bowler; dlt infers the schema, types the columns, and loads with the same governed
`merge` behaviour as batting. Snapshots replay just like batting: load v1, then a later export
merges on top and bowlers' running totals update in place.

Note: loaded **raw** (e.g. '-' left in STRIKE RATE / AVERAGE for wicketless bowlers). Cleaning
and derived metrics are dbt's job in Module 4 — the deliberate EL/T split.

    pip install openpyxl
    export DESTINATION__POSTGRES__CREDENTIALS="postgresql://app:<pw>@192.168.169.191:5432/appdb"
    python bowling_pipeline.py                       # loads bowling_2026_v1.xlsx
    python bowling_pipeline.py bowling_2026_v2.xlsx  # later export -> merges on top

Data lands in the `cricket` schema of `appdb` (table: bowling).
"""
import os
import sys
import dlt
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SNAP_DIR = os.environ.get("CRICKET_SNAPSHOT_DIR", os.path.join(HERE, "..", "data", "snapshots"))
SNAPSHOT = (sys.argv[1] if len(sys.argv) > 1
            else os.environ.get("CRICKET_BOWLING_SNAPSHOT", "bowling_2026_v1.xlsx"))
SEASON = int(os.environ.get("CRICKET_SEASON", "2026"))


@dlt.resource(name="bowling", write_disposition="merge", primary_key=("player", "season"))
def bowling(snapshot: str):
    """Read one bowling export (xlsx) and yield a row per bowler (header row -> dict keys).

    dlt normalizes the headers to safe column names ("ECONOMY RATE" -> economy_rate,
    "5 WICKET HAUL" -> _5_wicket_haul) and infers types. Wicketless bowlers carry '-' in
    STRIKE RATE / AVERAGE — dlt keeps the numbers in a typed column and parks '-' in a
    `..__v_text` variant column rather than erroring.
    """
    wb = openpyxl.load_workbook(os.path.join(SNAP_DIR, snapshot), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    for values in rows[1:]:
        if not values or values[1] in (None, "", "Player"):   # skip blanks; col 1 = Player
            continue
        record = dict(zip(header, values))
        record["season"] = SEASON
        yield record


def main() -> None:
    pipeline = dlt.pipeline(pipeline_name="cricket", destination="postgres", dataset_name="cricket")
    load_info = pipeline.run(bowling(SNAPSHOT))
    print(f"loaded snapshot: {SNAPSHOT}")
    print(load_info)
    print(f"  bowling: {len(pipeline.dataset()['bowling'].fetchall())} rows total in cricket.bowling")


if __name__ == "__main__":
    main()
