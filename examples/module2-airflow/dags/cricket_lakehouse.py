"""Day 27 + Day 48 (applied 🏏) — the cricket pipeline, orchestrated end-to-end in Airflow.

The whole platform in one scheduled DAG, running in-cluster:

    fetch ─dlt(merge)─► Postgres cricket.* ─[Soda gate]─► Trino CTAS ─► Iceberg lakehouse.cricket.*
                                                                              │
                                                                    dbt build (models + 39 tests)
                                                                              ▼
                                                                  lakehouse.cricket_dbt.* (marts)

- `load_to_postgres`  — dlt reads the batting/bowling/fielding snapshots (from the repo's raw
  URLs — the landed captures) and MERGEs them into Postgres `cricket.*`. Re-runs upsert, so
  running this weekly after each match keeps the tables current (Day 24's incremental, scheduled).
- `validate_landing` (Day 53) — a Soda Core scan of the Postgres landing zone (row counts, no
  duplicate players, valid economy rate, victims reconcile). Runs BEFORE promotion, so bad data
  never reaches Iceberg — quality as a pipeline stage. A failed check fails the task.
- `promote_to_lakehouse` — Trino materialises those Postgres tables as Iceberg tables in the
  lakehouse (Day 26's CTAS), so Trino/DuckDB/Spark all see the refreshed data.
- `transform_with_dbt` (Day 48) — runs the Module 4 dbt project (`examples/module4-dbt/
  cricket_lakehouse`) against Trino: builds the staging views + Iceberg marts into
  `lakehouse.cricket_dbt.*` AND runs the 39 tests. A failing test fails the task — so bad data
  never reaches the marts. This is the quality gate that turns dbt from "a thing you run" into
  part of the scheduled pipeline.

Secrets come from the Airflow connection `postgres_appdb` (not hard-coded). Trino needs no auth.
All tasks run as KubernetesExecutor pods that reach `pg-rw.postgres.svc`,
`trino.lakehouse.svc` and the internet directly.
"""
from __future__ import annotations

import pendulum
from airflow.decorators import dag, task

RAW = ("https://raw.githubusercontent.com/MichaelCade/90DaysOfDataEngineering/"
       "main/examples/cricket/data/snapshots")
SEASON = 2026


@dag(
    dag_id="cricket_lakehouse",
    schedule="0 7 * * 1",                 # 07:00 every Monday (post weekend fixtures)
    start_date=pendulum.datetime(2026, 8, 1, tz="UTC"),
    catchup=False,
    tags=["cricket", "dlt", "lakehouse", "dbt", "soda", "day27", "day48", "day53"],
)
def cricket_lakehouse():

    @task.virtualenv(requirements=["dlt[postgres]==1.28.0", "openpyxl", "requests"],
                     system_site_packages=True)
    def load_to_postgres():
        import io
        import os
        import json
        import dlt
        import requests
        import openpyxl
        from airflow.hooks.base import BaseHook

        c = BaseHook.get_connection("postgres_appdb")
        os.environ["DESTINATION__POSTGRES__CREDENTIALS"] = (
            f"postgresql://{c.login}:{c.password}@{c.host}:{c.port}/{c.schema}"
        )

        # NB: @task.virtualenv runs this in a fresh interpreter — module-level globals aren't
        # visible here, so define everything the nested resources close over *inside* the task.
        RAW = ("https://raw.githubusercontent.com/MichaelCade/90DaysOfDataEngineering/"
               "main/examples/cricket/data/snapshots")
        SEASON = 2026

        @dlt.resource(name="batting", write_disposition="merge", primary_key=("player", "season"))
        def batting():
            squad = requests.get(f"{RAW}/batting_2026_v2.json", timeout=30).json()
            for p in squad:
                if p.get("Total Runs", 0):
                    yield {"season": SEASON, **p}

        def _xlsx_rows(filename):
            wb = openpyxl.load_workbook(io.BytesIO(requests.get(f"{RAW}/{filename}", timeout=30).content),
                                        data_only=True)
            rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
            header = rows[0]
            for values in rows[1:]:
                if values and values[1] not in (None, "", "Player"):
                    yield {"season": SEASON, **dict(zip(header, values))}

        @dlt.resource(name="bowling", write_disposition="merge", primary_key=("player", "season"))
        def bowling():
            yield from _xlsx_rows("bowling_2026_v2.xlsx")

        @dlt.resource(name="fielding", write_disposition="merge", primary_key=("player", "season"))
        def fielding():
            yield from _xlsx_rows("fielding_2026_v1.xlsx")

        info = dlt.pipeline(pipeline_name="cricket", destination="postgres",
                            dataset_name="cricket").run([batting(), bowling(), fielding()])
        print(info)

    @task.virtualenv(requirements=["soda-core-postgres==3.5.6", "requests"],
                     system_site_packages=True)
    def validate_landing():
        # Day 53 — quality as a pipeline STAGE. Scan what dlt landed in Postgres BEFORE Trino
        # promotes it: if the landing data is bad, this task fails and promotion never runs, so
        # bad rows can't reach the Iceberg lakehouse. Creds come from the Airflow connection;
        # the SodaCL checks are the canonical file in the repo (single source of truth).
        import textwrap
        import requests
        from airflow.hooks.base import BaseHook
        from soda.scan import Scan

        c = BaseHook.get_connection("postgres_appdb")
        config = textwrap.dedent(f"""\
            data_source cricket_pg:
              type: postgres
              host: {c.host}
              port: {c.port}
              username: {c.login}
              password: {c.password}
              database: {c.schema}
              schema: cricket
        """)
        checks = requests.get(
            "https://raw.githubusercontent.com/MichaelCade/90DaysOfDataEngineering/"
            "main/examples/module5-quality/soda/checks/cricket.yml", timeout=30).text

        scan = Scan()
        scan.set_data_source_name("cricket_pg")
        scan.add_configuration_yaml_str(config)
        scan.add_sodacl_yaml_str(checks)
        scan.execute()
        print(scan.get_logs_text())
        scan.assert_no_checks_fail()          # raises -> task (and DAG run) fails on any bad check

    @task.virtualenv(requirements=["trino"], system_site_packages=True)
    def promote_to_lakehouse():
        from trino.dbapi import connect
        cur = connect(host="trino.lakehouse.svc.cluster.local", port=8080, user="airflow").cursor()
        cur.execute("CREATE SCHEMA IF NOT EXISTS lakehouse.cricket"); cur.fetchall()
        for t in ("batting", "bowling", "fielding"):
            cur.execute(f"DROP TABLE IF EXISTS lakehouse.cricket.{t}"); cur.fetchall()
            cur.execute(f"CREATE TABLE lakehouse.cricket.{t} AS SELECT * FROM postgres.cricket.{t}")
            cur.fetchall()
            cur.execute(f"SELECT count(*) FROM lakehouse.cricket.{t}")
            print(f"lakehouse.cricket.{t}:", cur.fetchall()[0][0], "rows")

    @task.virtualenv(requirements=["dbt-trino==1.10.3"], system_site_packages=True)
    def transform_with_dbt():
        # The dbt project isn't in the git-synced dags/ subPath, so fetch it from the public repo
        # (same "start at the file the repo already holds" idea the dlt task uses for snapshots),
        # then drive dbt in-process with dbtRunner — no dbt CLI on PATH to depend on.
        import os
        import subprocess
        import tempfile
        import textwrap
        from dbt.cli.main import dbtRunner

        work = tempfile.mkdtemp()
        repo = os.path.join(work, "repo")
        subprocess.run(
            ["git", "clone", "--depth", "1",
             "https://github.com/MichaelCade/90DaysOfDataEngineering.git", repo],
            check=True,
        )
        project = os.path.join(repo, "examples", "module4-dbt", "cricket_lakehouse")

        # In-cluster profile: point dbt at the Trino *service* (no auth), same host the CTAS uses.
        # (The committed profiles.yml targets the laptop LoadBalancer; here we want service DNS.)
        profiles_dir = os.path.join(work, "profiles")
        os.makedirs(profiles_dir, exist_ok=True)
        with open(os.path.join(profiles_dir, "profiles.yml"), "w") as f:
            f.write(textwrap.dedent("""\
                cricket_lakehouse:
                  target: prod
                  outputs:
                    prod:
                      type: trino
                      method: none
                      user: airflow
                      host: trino.lakehouse.svc.cluster.local
                      port: 8080
                      http_scheme: http
                      catalog: lakehouse
                      schema: cricket_dbt
                      threads: 4
            """))

        runner = dbtRunner()
        for cmd in (["deps"], ["build"]):
            res = runner.invoke([*cmd, "--project-dir", project, "--profiles-dir", profiles_dir])
            if not res.success:
                raise RuntimeError(f"dbt {cmd[0]} failed: {res.exception or 'see task logs'}")
        print("dbt build OK — lakehouse.cricket_dbt.* rebuilt and all tests passed")

    (load_to_postgres() >> validate_landing()
     >> promote_to_lakehouse() >> transform_with_dbt())


cricket_lakehouse()
